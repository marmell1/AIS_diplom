# from _4_Processing import get_input as get_input
from dataclasses import replace
from datetime import datetime

from _4_Processing import handler as handler
from database import get_session
from models import Products,WarehousesLoading
from sqlalchemy import select
import math
import pandas as pd
import getpass


import openpyxl
from openpyxl.styles import (Alignment, Font)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

import typer
app = typer.Typer()

@app.command()
def move (wh: int = typer.Option(..., help="Склад, на который перемещаем либо покупатель"),
          pr: int = typer.Option(..., help="Товар, который перемещаем"),
          q: int = typer.Option(..., help="Количество товара"),
          source: int = typer.Option(..., help="Склад, с которого списываем либо поставщик")):
    """
    Добавление товара на склад - ввести номер склада, id продукта, количество, склад-источник
    """
    #создать если нет такого продукта

    username = getpass.getuser()

    dict_serv = {"input_type": "CLI",
                 "resp": username,
                 "wh_in": wh,
                 "wh_out": source,
                 "product_id": pr,
                 "count": q
                 }

    # dict_serv = get_input("CLI", username, wh,pr,q,source)
    handler(dict_serv)

    print(f"Внесли склад номер {wh}, товар {pr}, {q} ед., источник - {source}")

def check_temp(rr):
    if rr == "стандартные":
        temp = 20
    elif rr == "не выше 10 градусов":
        temp = 0
    elif rr == "заморозка":
        temp = -20
    else:
        temp = "нет данных"
    return temp

@app.command()
def read(filename, wh_id):
    #инструкцию прописать


    """
    Загрузка данных из накладной поставщика. Для примера всего два поставщика - ИП Ромашка и ООО Подворье
    """
    workbook = openpyxl.load_workbook(filename, data_only=True)  # data_only=True считывает значения, а не формулы
    sheet = workbook.active
    # проверить что такого номера накладной нет в реестре
    # проверить что файл по правильной форме
    # что поставщик в списке
    # что суммы и цены неотрицательные
    # что дата еще не случилась
    # что все штуки ниже где надо integer


    with get_session() as session1:

        dic_vendors = {"ИП Ромашка":1,"ООО Подворье":2}
        dic_bill_check_vendor_row={1:14,2:1}
        dic_bill_check_vendor_column={1:9,2:1}
        dic_bill_number_row={1:26,2:3}
        dic_bill_number_column={1:50,2:2}
        dic_bill_number_date_row={1:26,2:2}
        dic_bill_number_date_column={1:61,2:1}
        dic_bill_first_row={1:30,2:5}
        dic_bill_sku_column={1:19,2:0}
        dic_bill_name_column={1:3,2:1}
        dic_bill_products_quantity_column={1:53,2:5}
        dic_bill_producer_column={1:48,2:2}
        dic_bill_measure_column={1:23,2:4}
        dic_bill_price_column1={1:59,2:6}
        dic_bill_price_column2={2:5}
        dic_bill_temp_column={1:33,2:3}
        dic_bill_capacity_column1={1:38,2:7}
        dic_bill_capacity_column2={2:5}

        for key in dic_vendors.keys():
            if sheet.cell(row=dic_bill_check_vendor_row[dic_vendors[key]], column=dic_bill_check_vendor_column[dic_vendors[key]]).value == key:
                act_number = sheet.cell(row=dic_bill_number_row[dic_vendors[key]], column=dic_bill_number_column[dic_vendors[key]]).value
                act_date = sheet.cell(row=dic_bill_number_date_row[dic_vendors[key]], column=dic_bill_number_date_column[dic_vendors[key]]).value
                input_type_name_date = "Накладная " + str(act_number) + " от " + str(act_date) + ", " + key
                print(act_number, act_date)

                rows = list(sheet.iter_rows(values_only=True))
                for row in rows[dic_bill_first_row[dic_vendors[key]]:]:
                    if row[0] is not None:
                        stmt = select(Products.id).where(Products.SKU == row[dic_bill_sku_column[dic_vendors[key]]])
                        pr_id = session1.scalar(stmt)
                        if pr_id is None:
                            temp = check_temp(row[dic_bill_temp_column[dic_vendors[key]]])
                            if dic_vendors[key] ==1:
                                price = row[dic_bill_price_column1[dic_vendors[key]]]
                                capacity = int(math.ceil(12 / row[dic_bill_capacity_column1[dic_vendors[key]]]))

                            if dic_vendors[key] ==2:
                                price = int(round(row[dic_bill_price_column1[dic_vendors[key]]]/row[dic_bill_price_column2[dic_vendors[key]]]))
                                capacity = int(math.ceil(row[dic_bill_capacity_column1[dic_vendors[key]]] / row[dic_bill_capacity_column2[dic_vendors[key]]]))

                            product1 = Products(SKU=row[dic_bill_sku_column[dic_vendors[key]]],
                                                name=row[dic_bill_name_column[dic_vendors[key]]],
                                                producer=row[dic_bill_producer_column[dic_vendors[key]]],
                                                measure=row[dic_bill_measure_column[dic_vendors[key]]],
                                                price=price,
                                                temp=temp,
                                                capacity=capacity)
                            session1.add(product1)
                            session1.commit()
                            stmt = select(Products.id).where(Products.SKU == row[dic_bill_sku_column[dic_vendors[key]]])
                            pr_id = session1.scalar(stmt)
                        # dict_serv = get_input("Report", input_type_name_date,wh_id, pr_id,
                        #                       int(row[dic_bill_products_quantity_column[dic_vendors[key]]]), "ИП Ромашка")

                        dict_serv = {"input_type":"Report",
                                     "resp":input_type_name_date,
                                     "wh_in":wh_id,
                                     "wh_out":key,
                                     "product_id":pr_id,
                                     "count":int(row[dic_bill_products_quantity_column[dic_vendors[key]]]),
                                     "date":act_date
                                     }

                        handler(dict_serv)
                    else:
                        break




        #
        # if sheet.cell(row=14, column=9).value == "ИП Ромашка":
        #
        #     act_number = sheet.cell(row=26, column=50).value
        #     act_date = sheet.cell(row=26, column=61).value
        #     input_type_name_date = "Накладная " + act_number + " от " + act_date
        #     print(act_number, act_date)
        #
        #     rows = list(sheet.iter_rows(values_only=True))
        #     for row in rows[30:]:
        #         if row[0] is not None:
        #             stmt = select(Products.id).where(Products.SKU == row[19])
        #             pr_id = session1.scalar(stmt)
        #             if pr_id is None:
        #                 temp = check_temp(row[33])
        #                 product1 = Products(SKU=row[19], name=row[3], producer=row[48],
        #                                     measure=row[23], price=row[59], temp=temp,
        #                                     capacity=int(math.ceil(12/row[38])));
        #                 session1.add(product1)
        #                 session1.commit()
        #                 stmt = select(Products.id).where(Products.SKU == row[19])
        #                 pr_id = session1.scalar(stmt)
        #             dict_serv = get_input("Report", wh_id, pr_id, int(row[53]), "ИП Ромашка")
        #             handler(dict_serv)
        #         else:
        #             break
        #
        # if sheet.cell(row=1, column=1).value == "ООО Подворье":
        #     act_number = sheet.cell(row=3, column=2).value
        #     act_date = sheet.cell(row=2, column=1).value
        #     input_type_name_date = "Накладная " + act_number + " от " + act_date
        #     print(act_number, act_date)
        #
        #     rows = list(sheet.iter_rows(values_only=True))
        #     for row in rows[5:]:
        #         if row[0] is not None:
        #             stmt = select(Products.id).where(Products.SKU == row[0])
        #             pr_id = session1.scalar(stmt)
        #             if pr_id is None:
        #                 temp = check_temp(row[3])
        #                 product1 = Products(SKU=row[0], name=row[1], producer=row[2],
        #                                     measure=row[4], price=int(round(row[6]/row[5])), temp=temp,
        #                                     capacity=int(math.ceil(row[7]/row[5])));
        #                 session1.add(product1)
        #                 session1.commit()
        #                 stmt = select(Products.id).where(Products.SKU == row[0])
        #                 pr_id = session1.scalar(stmt)
        #             dict_serv = get_input("Report", wh_id, pr_id, int(row[5]), "ООО Подворье")
        #             handler(dict_serv)
        #         else:
        #             break

@app.command()
def print_rep (type_rep: str = typer.Option(..., help="Вид отчета - Ведомость по остаткам на складах (balance) или Акт инвентаризации данных (inventory)"),
           wh_id: int = typer.Option(..., help="Номер склада"),
           filename: str = typer.Option("..def", help="Название файла, в который будут выгружены данные")):
    #инструкцию прописать
    """
    inventory balance report
    Генерация документов: ведомость по остаткам на складах и акт инвентаризации по конкретному складу.

    """
    dict_capt = {"balance":"Остатки на складе ","inventory":"Акт инвертаризации по складу "}
    today = datetime.now()
    now_date =  str(today.date()) + "_"+str(today.time()).replace(":","-")[:-7]
    if filename == "..def":
        filename = dict_capt[type_rep] + str(wh_id) +" на " + now_date + ".xlsx"
    elif ".xlsx" not in filename:
        filename = filename + ".xlsx"

    with get_session() as session1:
        stmt = select(WarehousesLoading).where(WarehousesLoading.warehouse_id == wh_id)
        db_data = session1.scalars(stmt).all()
        data_list = [
            {k: v for k, v in item.__dict__.items() if k != "_sa_instance_state"}
            for item in db_data
        ]
        df_whl = pd.DataFrame(data_list)

        stmt = select(Products)
        db_data = session1.scalars(stmt).all()
        data_list = [
            {k: v for k, v in item.__dict__.items() if k != "_sa_instance_state"}
            for item in db_data
        ]
        df_products = pd.DataFrame(data_list)

        df = df_whl.merge(df_products,left_on = 'product_id', right_on = "id", how = 'left')
        df['cost'] = df['quantity']*df['price']
        print(df.columns)
        wh_name = df['warehouse_name'][0]

        def format_capture(start_row):
            for col_num in range(1, len(df.columns) + 1):
                print(col_num)
                cell = worksheet.cell(row=start_row, column=col_num)
                cell.font = Font(bold=True)
                col_letter = get_column_letter(col_num)
                worksheet.column_dimensions[col_letter].width = len(df.columns[col_num - 1]) + 5

        final_row = 8 + len(df)

        if type_rep == "balance":
            df = df[['SKU', 'product_name', 'producer', 'quantity', 'cost']]
            df.columns = ['Артикул', 'Наименование товара', 'Изготовитель', 'Текущий остаток',
                          'Общая стоимость позиции на складе']
            start_row = 5 #сдвиг вниз
            with pd.ExcelWriter(filename, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, startrow=start_row)
                worksheet = writer.sheets["Sheet1"]  # Название листа по умолчанию
                worksheet["A1"] = "Ведомость остатков товаров"
                worksheet["A1"].font = Font(bold = True)
                worksheet["A2"] = f"Склад {wh_id} ({wh_name})"
                worksheet["A3"] = "Выгружено " + now_date


                format_capture(start_row+1)


                worksheet["A"+ str(final_row)] = "Итого"
                worksheet["A"+ str(final_row)].font = Font(bold = True)
                worksheet["E"+ str(final_row)] = df['Общая стоимость позиции на складе'].sum()
                worksheet["E"+ str(final_row)].font = Font(bold = True)

                for row in range(5,final_row+3):
                    worksheet["E"+ str(row)].number_format = "#,##0.00"


        elif type_rep == "inventory":
            df["Фактическое количество"] = ""
            df = df[['SKU', 'product_name', 'quantity', "Фактическое количество"]]
            df.columns = ['Артикул', 'Наименование товара', 'Учетное количество', "Фактическое количество"]
            start_row = 5 #сдвиг вниз
            df = df.sort_values(by='Учетное количество', ascending=False)

            with pd.ExcelWriter(filename, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, startrow=5)
                worksheet = writer.sheets["Sheet1"]  # Название листа по умолчанию
                worksheet["A1"] = "Акт инвентаризации"
                worksheet["A1"].font = Font(bold = True)
                worksheet["A2"] = f"Склад {wh_id} ({wh_name})"
                worksheet["A3"] = "Дата проверки -  " + now_date

                format_capture(start_row+1)

                chart = BarChart()
                chart.type = "bar"
                chart.style = 10  # Стиль оформления Excel (от 1 до 48)
                chart.title = "5 наиболее объемных товаров"
                chart.y_axis.title = "Количество"
                chart.x_axis.title = "Продукты"

                # 1. Указываем, откуда брать ЧИСЛА для столбиков (Колонка B, строки с 6 по финальную)
                # min_col=2 (это колонка B), min_row=6 (шапка), max_row=final_row
                data_ref = Reference(
                    worksheet, min_col=3, min_row=6, max_row=11
                )  # Передаем вместе с шапкой

                # 2. Указываем, откуда брать ТЕКСТ для подписей снизу (Колонка A, строки с 7 по финальную)
                cats_ref = Reference(worksheet, min_col=2, min_row=7, max_row=11)

                # Добавляем данные в график
                chart.add_data(data_ref, titles_from_data=True)  # titles_from_data подхватит имя из шапки
                chart.set_categories(cats_ref)

                # 3. Помещаем график на лист Excel (например, в ячейку D6, чтобы не перекрывать таблицу)
                worksheet.add_chart(chart, "G6")





        else:
            print()
            print()
            print ("Введите  корректное название отчета - balance (Ведомость по остаткам на складах) или inventory (Акт инвентаризации данных)"),
            print()
            print()

if __name__ == "__main__":
    app()